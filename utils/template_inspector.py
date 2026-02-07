from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


HEADER_ROWS = (4, 5)  # merged header spans these rows


def get_cell_text(ws, row, col):
    """
    Safely get text from merged or normal cells.
    """
    cell = ws.cell(row=row, column=col)

    # If part of merged cell, read from top-left
    if cell.coordinate in ws.merged_cells:
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(
                    row=merged.min_row,
                    column=merged.min_col
                ).value
    return cell.value


def find_columns(ws):
    total_candidates = []
    percentage_candidates = []

    for col in range(1, ws.max_column + 1):
        texts = []

        for row in HEADER_ROWS:
            value = get_cell_text(ws, row, col)
            if value:
                texts.append(str(value).lower())

        combined_text = " ".join(texts)

        if "total" in combined_text:
            total_candidates.append(col)

        if "percentage" in combined_text or "%" in combined_text:
            percentage_candidates.append(col)

    total_col = max(total_candidates) if total_candidates else None
    percentage_col = max(percentage_candidates) if percentage_candidates else None

    return total_col, percentage_col

def inspect_templates(template_dir: Path):
    for template in template_dir.glob("*.xlsx"):
        wb = load_workbook(template)
        ws = wb.active  # or explicit sheet if you want

        total_col, percentage_col = find_columns(ws)

        print(f"\nTemplate: {template.name}")

        if total_col:
            print(f"  Grand Total     : {get_column_letter(total_col)} ({total_col})")
        else:
            print("  Grand Total     : NOT FOUND")

        if percentage_col:
            print(f"  Percentage      : {get_column_letter(percentage_col)} ({percentage_col})")
        else:
            print("  Percentage      : NOT FOUND")


if __name__ == "__main__":
    TEMPLATE_DIR = Path("backend/templates")
    inspect_templates(TEMPLATE_DIR)
