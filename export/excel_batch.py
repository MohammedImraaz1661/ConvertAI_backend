from export.template_loader import load_template
from openpyxl.styles import Font, PatternFill
from export.writer import write_student_result, TOPPER_FILL


def write_batch_results_excel(
    results: list[dict],
    template_path: str,
    output_path: str,
    start_row: int = 6
):
    wb, ws, subject_column_map, total_col, percentage_col = load_template(template_path)

    row = start_row
    sl_no = 1
    topper_tracker = []

    for result in results:
        wb, numeric_total = write_student_result(
            wb=wb,
            ws=ws,
            subject_column_map=subject_column_map,
            result_data={
                "header": result["header"],
                "subjects": result["subjects"]
            },
            row_index=row,
            sl_no=sl_no,
            total_sum_col=total_col,
            percentage_col=percentage_col
        )

        topper_tracker.append((row, numeric_total))
        row += 1
        sl_no += 1

    # 🏆 Highlight class topper
    if topper_tracker:
        topper_row, _ = max(topper_tracker, key=lambda x: x[1])
        for col in range(1, ws.max_column + 1):
            ws.cell(row=topper_row, column=col).fill = TOPPER_FILL

    # 🔤 Bold important columns
    bold_font = Font(bold=True)
    BOLD_COLUMNS = {1, 2, 3, total_col, percentage_col}

    for cols in subject_column_map.values():
        if "TOTAL" in cols:
            BOLD_COLUMNS.add(cols["TOTAL"])

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.column in BOLD_COLUMNS:
                cell.font = bold_font

    # 🗂 Legend
    LEGEND_FAIL = PatternFill(start_color="FFD60A", end_color="FFD60A", fill_type="solid")
    LEGEND_PE = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    LEGEND_NSS = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
    LEGEND_TOPPER = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

    legend_col = ws.max_column + 2
    legend_row = start_row

    title = ws.cell(row=legend_row, column=legend_col)
    title.value = "LEGEND"
    title.font = Font(bold=True)

    ws.cell(row=legend_row + 2, column=legend_col, value="Subject Fail").fill = LEGEND_FAIL
    ws.cell(row=legend_row + 3, column=legend_col, value="(BPEK559)").fill = LEGEND_PE
    ws.cell(row=legend_row + 4, column=legend_col, value="(BNSK559)").fill = LEGEND_NSS
    ws.cell(row=legend_row + 5, column=legend_col, value="Class Topper").fill = LEGEND_TOPPER

    wb.save(output_path)
