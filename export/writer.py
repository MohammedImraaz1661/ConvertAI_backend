from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill
import re

# Colours
FAIL_FILL = PatternFill(start_color="FFD60A", end_color="FFD60A", fill_type="solid")
PE_FILL = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
NSS_FILL = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
TOPPER_FILL = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

IGNORE_FAIL_COLOR_SUBJECTS = {"BAI586", "BAI786"}
NO_EXTERNAL_FAIL_SUBJECTS = set()
MAJOR_PROJECT_SUBJECTS = {"BAI786"}
MINI_PROJECT_SUBJECTS = {"BAI586"}

def is_activity_subject(code: str) -> bool:
    return bool(code) and code[-1] == "9"


def subject_sort_key(code: str):
    if not code:
        return ("", 0, "")
    match = re.match(r"([A-Z]+)(\d+)([A-Z]?)", code)
    if not match:
        return (code, 0, "")
    prefix, number, suffix = match.groups()
    return (prefix, int(number), suffix)


def write_student_result(
    wb: Workbook,
    ws,
    subject_column_map: dict,
    result_data: dict,
    row_index: int,
    sl_no: int,
    total_sum_col: int | None,
    percentage_col: int | None,
):
    header = result_data["header"]
    subjects = sorted(
        result_data["subjects"],
        key=lambda s: subject_sort_key(s.get("subject_code", ""))
    )

    for s in subjects:
        print(" ", s["subject_code"], s["internal"], s["external"], s["total"])

    # print("\nSUBJECT COLUMN MAP KEYS:")
    for k in subject_column_map:
        print(" ", k)

    # Student info
    ws.cell(row=row_index, column=1).value = sl_no
    ws.cell(row=row_index, column=2).value = header["usn"]
    ws.cell(row=row_index, column=3).value = header["name"]

    numeric_total = 0
    max_total = 0
    activity_subjects = []

    # Detect activity slots FIRST
    activity_keys = [k for k in subject_column_map if "/" in k]
    has_activity_slots = bool(activity_keys)

    # print("\nACTIVITY SLOTS IN TEMPLATE:", activity_keys)
    # print("HAS ACTIVITY SLOTS:", has_activity_slots)

    # ---------- MAIN SUBJECT LOOP ----------
    for subject in subjects:
        code = subject.get("subject_code")
        # print("\nPROCESSING SUBJECT:", code)

        if not code:
            # print("  ⛔ skipped: empty code")
            continue

        if is_activity_subject(code) and has_activity_slots:
            # print("  ➡ routed as ACTIVITY subject")
            activity_subjects.append(subject)
            continue

        if code not in subject_column_map:
            # print("  ⛔ skipped: code not in subject_column_map")
            continue

        cols = subject_column_map[code]
        # print("  ✅ writing to columns:", cols)

        int_marks = subject["internal"]
        ext_marks = subject["external"]
        tot_marks = subject["total"]

        int_cell = ws.cell(row=row_index, column=cols["INTERNAL"])
        ext_cell = ws.cell(row=row_index, column=cols["EXTERNAL"])
        tot_cell = ws.cell(row=row_index, column=cols["TOTAL"])

        int_cell.value = int_marks
        ext_cell.value = ext_marks
        tot_cell.value = tot_marks

        fail = False
        if code not in IGNORE_FAIL_COLOR_SUBJECTS:
            fail = int_marks < 18 or ext_marks < 18 or tot_marks < 36

        if fail:
            # print("  ❌ FAIL → marking yellow")
            int_cell.fill = FAIL_FILL
            ext_cell.fill = FAIL_FILL
            tot_cell.fill = FAIL_FILL

        numeric_total += tot_marks
        if code in MAJOR_PROJECT_SUBJECTS:
            max_total+=200
        else:
            max_total+=100

        # print("  ✔ added to total:", tot_marks)
        # print("  ✔ max_total now:", max_total)

    # ---------- ACTIVITY SUBJECTS ----------
    # print("\nACTIVITY SUBJECTS COLLECTED:", [s["subject_code"] for s in activity_subjects])

    for activity_subject, activity_key in zip(activity_subjects, activity_keys):
        # print("WRITING ACTIVITY:", activity_subject["subject_code"], "→", activity_key)

        cols = subject_column_map[activity_key]

        int_cell = ws.cell(row=row_index, column=cols["INTERNAL"])
        ext_cell = ws.cell(row=row_index, column=cols["EXTERNAL"])
        tot_cell = ws.cell(row=row_index, column=cols["TOTAL"])

        int_cell.value = activity_subject["internal"]
        ext_cell.value = activity_subject["external"]
        tot_cell.value = activity_subject["total"]

        if activity_subject["subject_code"] == "BPEK559":
            fill = PE_FILL
        elif activity_subject["subject_code"] == "BNSK559":
            fill = NSS_FILL
        else:
            fill = None

        if fill:
            int_cell.fill = fill
            ext_cell.fill = fill
            tot_cell.fill = fill

        numeric_total += activity_subject["total"]
        max_total += 100

    # print("\nFINAL TOTAL:", numeric_total)
    # print("FINAL MAX:", max_total)

    # ---------- GRAND TOTAL ----------
    if total_sum_col:
        ws.cell(row=row_index, column=total_sum_col).value = numeric_total

    # ---------- PERCENTAGE ----------
    if percentage_col and max_total > 0:
        percentage = round((numeric_total / max_total) * 100, 1)
        ws.cell(row=row_index, column=percentage_col).value = percentage
        # print("FINAL %:", percentage)

    # print("==============================\n")

    return wb, numeric_total
