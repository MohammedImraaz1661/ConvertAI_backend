from openpyxl import load_workbook
import re

SUBJECT_CODE_PATTERN = re.compile(r"^[A-Z]{3,5}\d{3,4}[A-Z]?$")

HEADER_ROWS = (4, 5)  # merged header rows


def is_activity_header(text: str) -> bool:
    if "/" not in text:
        return False
    parts = [p.strip() for p in text.split("/")]
    return (
        len(parts) == 2
        and all(SUBJECT_CODE_PATTERN.match(p) and p[-1] == "9" for p in parts)
    )


def get_cell_text(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.coordinate in ws.merged_cells:
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(
                    row=merged.min_row,
                    column=merged.min_col
                ).value
    return cell.value


def find_total_and_percentage_columns(ws):
    total_col = None
    percentage_col = None

    for merged in ws.merged_cells.ranges:
        # Only care about header rows
        if merged.min_row not in (4, 5):
            continue

        header_cell = ws.cell(row=merged.min_row, column=merged.min_col)
        if not header_cell.value:
            continue

        text = str(header_cell.value).lower()

        # VTU puts numeric data at the END of merged header
        data_col = merged.max_col

        if "total" in text:
            total_col = data_col

        if "percentage" in text or "%" in text:
            percentage_col = data_col

    return total_col, percentage_col


def load_template(template_path: str):
    wb = load_workbook(template_path)

    if "Student Result" in wb.sheetnames:
        ws = wb["Student Result"]
    elif "Student Results" in wb.sheetnames:
        ws = wb["Student Results"]
    else:
        ws = wb.worksheets[0]

    subject_column_map = {}
    col = 1
    max_col = ws.max_column

    while col <= max_col:
        header_value = None

        for r in HEADER_ROWS:
            val = get_cell_text(ws, r, col)
            if isinstance(val, str):
                header_value = val.strip()
                break

        if header_value:
            if is_activity_header(header_value):
                subject_column_map[header_value] = {
                    "INTERNAL": col,
                    "EXTERNAL": col + 1,
                    "TOTAL": col + 2,
                }
                col += 3
                continue

            if SUBJECT_CODE_PATTERN.match(header_value):
                subject_column_map[header_value] = {
                    "INTERNAL": col,
                    "EXTERNAL": col + 1,
                    "TOTAL": col + 2,
                }
                col += 3
                continue

        col += 1

    total_col, percentage_col = find_total_and_percentage_columns(ws)
    # print("TOTAL COL:", total_col)
    # print("PERCENT COL:", percentage_col)


    return wb, ws, subject_column_map, total_col, percentage_col
